"""
使用openpyxl实现以下需求

使用excel 写入一组数据，姓名，身高，体重
计算是否为健康体重，如果是健康体重，则在旁边备注健康，并将姓名打印出来
健康体重计算公式：（身高cm-70）×60%
(可以做一部分优化)
"""

from openpyxl import Workbook, load_workbook

class PracticeExcel:  # 初始化一个PracticeExcel类

    wb = Workbook()  # 定义一个变量wb
    ws1 = wb.active  # 激活当前页面
    ws1.title = "create"  # 给此页面赋值
    ws1["A1"] = "姓名"  # 给A1单元格赋值
    ws1["B1"] = "身高"  # 给B1单元格赋值
    ws1["C1"] = "体重"  # 给C1单元格赋值
    ws1["D1"] = "备注"  # 给D1单元格赋值
    name = ["dong", "fang", "shen", "qi"]  # 定义一个列表
    data_dict1 = {"dong": 180, "fang": 160, "shen": 170, "qi": 155}  # 定义一个字典，用列表的值做为字典的key值
    data_keys1 = [i for i in data_dict1.keys()]  # 取出字典的key值
    data_dict2 = {data_dict1[data_keys1[0]]: 66, data_dict1[data_keys1[1]]: 51, data_dict1[data_keys1[2]]: 60, data_dict1[data_keys1[3]]: 81}
    # 用上个字典的value值做为这个字典的key值
    data_keys2 = [i for i in data_dict2.keys()]  # 取出字典的key值

    for i in range(len(name)):  # 当i在列表长度范围内，做循环
        ws1.cell(row=i + 2, column=1).value = data_keys1[i]  # 将data_keys1[i]的值循环放入列表第一列的第i+2行中
        ws1.cell(row=i + 2, column=2).value = data_keys2[i]  # 将data_keys2[i]的值循环放入列表第二列的第i+2行中
        ws1.cell(row=i + 2, column=3).value = data_dict2[data_keys2[i]]  # 将data_dict2[data_keys2[i]]的值循环放入列表第三列的第i+2行中
    wb.save("data.xlsx")  # 保存Excel表以data名字，xlsx格式

    def health_data(self):  # 定义一个health_data方法
        ld = load_workbook(filename="data.xlsx")  # 取出data.xlsx文件
        sheet = ld["create"]  # 打开名称为create的页面

        for i in range(4):  # 循环4次
            height = sheet.cell(row=i + 2, column=2).value  # 取出第2列第i+2行的数据
            weight = sheet.cell(row=i + 2, column=3).value  # 取出第3列第i+2行的数据
            name = sheet.cell(row=i + 2, column=1).value  # 取出第1列第i+2行的数据

            health_weight = (height - 70) * 0.6  # 定义一个健康体重的标准
            if weight == health_weight:  # 判断当weight等于健康体重标准时
                print(name, "这是健康的体重", weight)  # 打印出健康体重
                sheet.cell(row=i + 2, column=4).value = "健康体重"  # 在第4列的i+2位置写入“健康体重”
        ld.save(filename="data.xlsx")  # 保存data.xlsx表

pe = PracticeExcel()  # 实例化类
pe.health_data()  # 调用health_data方法
